"""Turn a raw Qualtrics export into the cleaned, coded table."""
from __future__ import annotations

from dataclasses import dataclass, field

import openpyxl

from .profile import parse_numeric
from .reference import CityReference


@dataclass
class CleanedColumn:
    header: str          # question text shown in row 1
    kind: str            # coded | numeric | multi_select_option | free_text
    values: list         # one entry per respondent
    codes: dict = field(default_factory=dict)   # label -> code, for crosstab options
    source_qid: str = ""
    option_label: str = ""                      # set for multi-select indicators


@dataclass
class CleanResult:
    columns: list[CleanedColumn]
    cities: list[str]
    populations: list
    quintiles: list
    regions: list
    issues: list[str]
    unmatched: list[str]
    fuzzy: list[tuple]

    @property
    def n(self) -> int:
        return len(self.cities)


def clean(path, specs, reference: CityReference, sheet, first_data_row, last_data_row,
          city_overrides: dict | None = None):
    """Build the cleaned table. `city_overrides` maps a raw survey-entered name
    to a confirmed reference city, for names the matcher could not resolve alone."""
    overrides = {k.strip().lower(): v for k, v in (city_overrides or {}).items()}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]

    city_spec = next((s for s in specs if s.role == "city"), None)
    if city_spec is None:
        raise ValueError("no city column identified -- set one column's role to 'city'")

    issues, unmatched, fuzzy = [], [], []
    keep_rows, cities, pops, quints, regs = [], [], [], [], []

    for r in range(first_data_row, last_data_row + 1):
        raw_city = ws.cell(r, city_spec.index).value
        if raw_city in (None, ""):
            issues.append(f"row {r}: blank city -- row dropped")
            continue

        if str(raw_city).strip().lower() in overrides:
            city = reference.cities.get(overrides[str(raw_city).strip().lower()])
            conf = "override"
        else:
            city, conf = reference.match(raw_city)

        if city is None:
            unmatched.append(str(raw_city))
            issues.append(f"row {r}: city {raw_city!r} not in reference -- row dropped")
            continue
        if conf == "fuzzy":
            fuzzy.append((str(raw_city), city.name))

        q = reference.quintile(city)
        if q is None or city.region_code is None:
            issues.append(f"row {r}: {city.name} missing population or region -- row dropped")
            continue

        keep_rows.append(r)
        cities.append(city.name)
        pops.append(city.population)
        quints.append(q)
        regs.append(city.region_code)

    # Flag duplicate submissions rather than silently keeping both.
    seen = {}
    for i, c in enumerate(cities):
        seen.setdefault(c, []).append(i)
    for c, idxs in seen.items():
        if len(idxs) > 1:
            issues.append(f"{c}: {len(idxs)} submissions -- all retained, review for duplicates")

    columns = []
    for spec in specs:
        if spec.role != "question" or spec.kind == "skip":
            continue
        raw_vals = [ws.cell(r, spec.index).value for r in keep_rows]

        if spec.kind == "coded":
            lookup = {str(k).strip().lower(): v for k, v in spec.codes.items()}
            vals, unknown = [], set()
            for v in raw_vals:
                s = str(v).strip() if v is not None else ""
                if not s:
                    vals.append(None)
                elif s.lower() in lookup:
                    vals.append(lookup[s.lower()])
                else:
                    vals.append(None)
                    unknown.add(s)
            if unknown:
                issues.append(
                    f"{spec.qid}: {len(unknown)} answer(s) not in codebook, left blank: "
                    + "; ".join(sorted(unknown)[:5])
                )
            columns.append(CleanedColumn(spec.text, "coded", vals, dict(spec.codes), spec.qid))

        elif spec.kind == "numeric":
            vals, blanked, annotated = [], [], []
            for i, v in enumerate(raw_vals):
                num, note = parse_numeric(v)
                vals.append(num)
                if num is None and v not in (None, ""):
                    blanked.append(f"{cities[i]}: {str(v)[:52]!r} ({note})")
                elif note:
                    annotated.append(f"{cities[i]}: {note}")
            # One summary line per question with a few examples, rather than one
            # line per answer -- a log nobody reads catches nothing.
            if blanked:
                issues.append(
                    f"{spec.qid} [{spec.text[:48]}]: {len(blanked)} answer(s) could not be "
                    f"read as a single number and were left blank. "
                    + " | ".join(blanked[:4])
                    + (f" | +{len(blanked) - 4} more" if len(blanked) > 4 else ""))
            if annotated:
                issues.append(
                    f"{spec.qid} [{spec.text[:48]}]: {len(annotated)} answer(s) had "
                    f"trailing text removed. " + " | ".join(annotated[:4])
                    + (f" | +{len(annotated) - 4} more" if len(annotated) > 4 else ""))

            # Only genuinely discontinuous values are flagged. A plain spread
            # rule fires on Portland and Salem in every population column, which
            # trains the reader to skip the log; a value orders of magnitude
            # above the next one down is the shape a units error actually takes.
            nums = sorted((v for v in vals if v is not None), reverse=True)
            if len(nums) >= 6 and nums[1] > 0 and nums[0] / nums[1] >= 25:
                worst = nums[0]
                who = [cities[i] for i, v in enumerate(vals) if v == worst]
                issues.append(
                    f"{spec.qid} [{spec.text[:38]}]: {', '.join(who[:3])} = {worst:,.6g}, "
                    f"more than 25x the next largest ({nums[1]:,.6g}) -- check units")
            columns.append(CleanedColumn(spec.text, "numeric", vals, {}, spec.qid))

        elif spec.kind == "multi_select":
            # One 0/1 indicator column per option. Exact token matching avoids the
            # substring collisions a wildcard COUNTIFS would produce (for example
            # "Late Fee" matching "Late Fee and Interest").
            for label in sorted(spec.codes, key=lambda k: spec.codes[k]):
                target = label.strip().lower()
                vals = []
                for v in raw_vals:
                    if v in (None, ""):
                        vals.append(None)
                        continue
                    toks = {p.strip().lower() for p in str(v).split(",") if p.strip()}
                    vals.append(1 if target in toks else 0)
                columns.append(CleanedColumn(
                    f"{spec.text} [{label}]", "multi_select_option", vals,
                    {"Selected": 1, "Not selected": 0}, spec.qid, option_label=label))

        elif spec.kind == "free_text":
            vals = ["" if v is None else str(v).strip() for v in raw_vals]
            columns.append(CleanedColumn(spec.text, "free_text", vals, {}, spec.qid))

    wb.close()
    return CleanResult(columns, cities, pops, quints, regs, issues, unmatched, fuzzy)
