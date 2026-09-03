"""Read a raw Qualtrics export and propose a codebook.

The codebook is the specification for both cleaning and analysis. It is written to
JSON so it can be reviewed and corrected before any workbook is generated -- the
response-code mapping is the one step that genuinely needs human judgement, and
getting it wrong is invisible in the finished spreadsheet.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass, asdict, field
from pathlib import Path

import openpyxl

# Qualtrics bookkeeping columns, dropped from the cleaned output.
METADATA_HEADERS = {
    "startdate", "enddate", "status", "ipaddress", "progress",
    "duration (in seconds)", "finished", "recordeddate", "responseid",
    "recipientlastname", "recipientfirstname", "recipientemail",
    "externalreference", "locationlatitude", "locationlongitude",
    "distributionchannel", "userlanguage",
}

# Respondent contact details: personally identifying, never published.
PII_PATTERNS = [
    r"your name", r"job title", r"email", r"phone",
    r"contact name", r"respondent name",
]

CITY_PATTERNS = [r"city name", r"name of (your )?city", r"^city$", r"municipality"]

# Answer sets encoded consistently across the existing workbooks.
KNOWN_SCALES = [
    (["yes", "no", "unsure"], {"yes": 1, "no": 2, "unsure": 3}),
    (["yes", "no"], {"yes": 1, "no": 2}),
    (["monthly", "bi-monthly", "quarterly"],
     {"monthly": 1, "bi-monthly": 2, "quarterly": 3}),
    (["yes", "no", "don't know"], {"yes": 1, "no": 2, "don't know": 3}),
]

NUMERIC_CLEAN = re.compile(r"[^0-9.\-]")


def looks_numeric(value) -> bool:
    """True when a free-text answer is really a number wearing punctuation."""
    if isinstance(value, (int, float)):
        return True
    s = str(value).strip()
    if not s:
        return False
    if re.fullmatch(r"[\$\s]*-?[\d,]+(\.\d+)?\s*%?", s):
        return True
    return False


# The decisive test is how many numbers the answer contains, not whether it is
# tidy. One number with a unit attached ("39 miles", "2,313 miles") is recoverable.
# Two or more is ambiguous and must not be guessed at: stripping the punctuation
# out of "$39 (20% pay $23/mnth, 60% pay $39/mnth)" and keeping the digits yields
# 39202360392056 -- a value that recalculates cleanly and is catastrophically wrong.
NUMBER_RE = re.compile(r"-?\$?\s*\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\$?\s*\d*\.?\d+")


def find_numbers(text: str):
    """Every distinct number in a string, as (value, matched_text) pairs."""
    out = []
    for m in NUMBER_RE.finditer(text):
        raw = m.group(0)
        cleaned = raw.replace("$", "").replace(",", "").strip()
        try:
            out.append((float(cleaned), raw.strip(), m.start(), m.end()))
        except ValueError:
            continue
    return out


def parse_numeric(value, allow_annotation: bool = True):
    """Return (number, note) or (None, reason).

    '$5.00' -> (5.0, '')            '39 miles' -> (39.0, "unit text 'miles' dropped")
    '1.5%'  -> (0.015, '')          '8/4'      -> (None, 'contains 2 numbers')
    """
    if isinstance(value, (int, float)):
        return float(value), ""
    if value is None:
        return None, "blank"
    s = str(value).strip()
    if not s:
        return None, "blank"

    nums = find_numbers(s)
    if not nums:
        return None, "no number present"
    if len(nums) > 1:
        return None, f"contains {len(nums)} numbers"

    val, raw, start, end = nums[0]
    remainder = (s[:start] + s[end:]).strip()
    # A percent sign anywhere converts to a fraction, matching the workbooks.
    if "%" in s:
        val = val / 100.0
        remainder = remainder.replace("%", "").strip()
    remainder = remainder.strip(" .,:;()[]-+")

    # "13 million" is 13,000,000, not 13. Left unscaled this is a 1,000,000x
    # error that looks entirely plausible sitting in a column of gallon figures.
    mag = re.match(r"^(million|billion|thousand|m|b|k)\b", remainder, re.IGNORECASE)
    if mag:
        val *= {"thousand": 1e3, "k": 1e3, "million": 1e6, "m": 1e6,
                "billion": 1e9, "b": 1e9}[mag.group(1).lower()]
        remainder = remainder[mag.end():].strip(" .,:;()[]-+")

    if remainder and not allow_annotation:
        return None, f"annotated: {remainder[:40]!r}"
    note = f"kept {val:g}, dropped {remainder[:35]!r}" if remainder else ""
    return val, note


@dataclass
class ColumnSpec:
    index: int                      # 1-based column in the raw sheet
    qid: str                        # Qualtrics question id, row 1
    text: str                       # human-readable question, row 2
    role: str                       # city | metadata | pii | question
    kind: str                       # coded | numeric | multi_select | free_text | skip
    codes: dict = field(default_factory=dict)   # answer label -> integer code
    note: str = ""

    def to_json(self):
        return asdict(self)


OTHER_RE = re.compile(r"^other\b|please specify", re.I)
MULTI_TEXT_RE = re.compile(r"check all that apply|select all that apply", re.I)


def _split_options(strs):
    """Comma-split answers into the underlying option set."""
    opts = []
    for s in strs:
        for part in s.split(","):
            p = part.strip()
            if p:
                opts.append(p)
    return opts


def _classify(values, header_text):
    """Decide how a column should be cleaned and analysed."""
    vals = [v for v in values if v not in (None, "")]
    if not vals:
        return "skip", {}, "no responses"

    strs = [str(v).strip() for v in vals]
    uniq = sorted({s for s in strs})
    is_text_col = bool(re.search(r"-\s*text\s*$", header_text, re.I))

    # Check-all-that-apply. The question wording is the reliable signal; the
    # value shape alone confuses these with single-select answers that happen
    # to contain a comma. The free-text "- Text" partner column is excluded.
    if MULTI_TEXT_RE.search(header_text) and not is_text_col:
        opts = _split_options(strs)
        distinct = sorted(set(opts))
        if len(distinct) <= 25:
            codes = {o: i + 1 for i, o in enumerate(distinct)}
            return "multi_select", codes, f"{len(distinct)} options, expanded to indicator columns"

    if all(looks_numeric(s) for s in strs):
        return "numeric", {}, ""

    # Known scales, tolerating an "Other (Please Specify)" tail option.
    core = [u for u in uniq if not OTHER_RE.match(u)]
    others = [u for u in uniq if OTHER_RE.match(u)]
    low = [c.lower() for c in core]
    for scale, mapping in KNOWN_SCALES:
        if core and set(low) <= set(scale):
            codes = {c: mapping[c.lower()] for c in core}
            nxt = max(codes.values()) + 1 if codes else 1
            for o in others:
                codes[o] = nxt
                nxt += 1
            note = "recognised answer scale"
            if others:
                note += " (+ Other)"
            return "coded", codes, note

    # Small closed set -> single-select. Codes are assigned alphabetically with
    # any "Other" pushed last, and MUST be reviewed: the numbers drive the
    # column order of every crosstab built from this question.
    if len(uniq) <= 12 and len(uniq) < len(strs) * 0.6 and not is_text_col:
        ordered = sorted(core) + sorted(others)
        codes = {u: i + 1 for i, u in enumerate(ordered)}
        return "coded", codes, "REVIEW: codes assigned alphabetically"

    # Mostly numbers with a few written-out answers ("1 day", "30 days").
    numlike = sum(1 for s in strs if parse_numeric(s)[0] is not None)
    if numlike / len(strs) > 0.7:
        return "numeric", {}, f"REVIEW: {len(strs) - numlike} non-numeric answers will be blanked"

    return "free_text", {}, "verbatim -- listed, not tabulated"


def profile_workbook(path, sheet=None, header_row=1, text_row=2, first_data_row=3):
    """Scan a raw export and return a list of ColumnSpec."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet is None:
        sheet = next((s for s in wb.sheetnames if s.upper() == "RAW"), wb.sheetnames[0])
    ws = wb[sheet]

    last_row = first_data_row - 1
    for r in range(first_data_row, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, min(ws.max_column, 40) + 1)):
            last_row = r

    specs = []
    for c in range(1, ws.max_column + 1):
        qid = str(ws.cell(header_row, c).value or "").strip()
        text = str(ws.cell(text_row, c).value or "").strip()
        blob = f"{qid} {text}".lower()
        values = [ws.cell(r, c).value for r in range(first_data_row, last_row + 1)]

        if qid.lower() in METADATA_HEADERS:
            specs.append(ColumnSpec(c, qid, text, "metadata", "skip", note="Qualtrics metadata"))
            continue
        if any(re.search(p, blob) for p in CITY_PATTERNS):
            specs.append(ColumnSpec(c, qid, text, "city", "skip", note="city identifier"))
            continue
        if any(re.search(p, blob) for p in PII_PATTERNS):
            specs.append(ColumnSpec(c, qid, text, "pii", "skip", note="respondent contact detail"))
            continue

        kind, codes, note = _classify(values, text)
        specs.append(ColumnSpec(c, qid, text, "question", kind, codes, note))

    wb.close()
    return specs, sheet, (first_data_row, last_row)


def write_codebook(specs, path, meta=None):
    payload = {
        "_instructions": [
            "Review every entry whose note begins with REVIEW.",
            "'codes' maps each answer label to the integer used in the analysis;",
            "edit the numbers to control the order options appear in the crosstabs.",
            "Set 'kind' to skip to drop a column from the output entirely.",
        ],
        "meta": meta or {},
        "columns": [s.to_json() for s in specs],
    }
    Path(path).write_text(json.dumps(payload, indent=2), encoding="utf-8")


def read_codebook(path):
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    return [ColumnSpec(**c) for c in payload["columns"]], payload.get("meta", {})
