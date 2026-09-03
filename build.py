"""Write the analysis workbook.

Every formula range in the output is derived from one pair of constants,
DATA_START and data_end. Nothing is typed twice, so the criteria-range
misalignment that affects the hand-built workbooks cannot occur here.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .reference import QUINTILE_LABELS, REGION_NAMES

DATA_START = 2
GROWTH_PAD = 40          # blank rows kept inside every range for next year's respondents

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
TOT_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

QUINTILE_COL = "C"
REGION_COL = "D"


def _style_header(cell):
    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    cell.fill = HDR_FILL
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _label(cell, text, bold=False, fill=None):
    cell.value = text
    cell.font = Font(name=FONT, bold=bold, size=10)
    if fill:
        cell.fill = fill
    cell.border = BOX


class AnalysisWorkbook:
    def __init__(self, result, reference, title="Survey Analysis"):
        self.res = result
        self.ref = reference
        self.title = title
        self.n = result.n
        self.data_end = DATA_START + self.n - 1 + GROWTH_PAD
        self.wb = Workbook()
        self.data_ws = self.wb.active
        self.data_ws.title = "Cleaned"
        self.col_of = {}          # index into result.columns -> sheet column letter

    # ---------------- cleaned data ----------------
    def write_data(self):
        ws = self.data_ws
        heads = ["CITY", "Population", "QCODE", "Region"] + [c.header for c in self.res.columns]
        for i, h in enumerate(heads, start=1):
            _style_header(ws.cell(1, i))
            ws.cell(1, i).value = h

        for r in range(self.n):
            row = DATA_START + r
            ws.cell(row, 1).value = self.res.cities[r]
            ws.cell(row, 2).value = self.res.populations[r]
            ws.cell(row, 3).value = self.res.quintiles[r]
            ws.cell(row, 4).value = self.res.regions[r]
            for ci, col in enumerate(self.res.columns):
                v = col.values[r]
                ws.cell(row, 5 + ci).value = v if v != "" else None

        for ci in range(len(self.res.columns)):
            self.col_of[ci] = get_column_letter(5 + ci)

        for c in range(1, len(heads) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 15 if c > 4 else 13
        ws.column_dimensions["A"].width = 20
        ws.row_dimensions[1].height = 58
        ws.freeze_panes = "E2"
        for row in ws.iter_rows(min_row=DATA_START, max_row=DATA_START + self.n - 1,
                                max_col=len(heads)):
            for cell in row:
                cell.font = Font(name=FONT, size=10)

    # ---------------- helpers ----------------
    # Analysis formulas live on a different sheet from the data, so every range
    # must name the Cleaned sheet. Unqualified, they silently read the Analysis
    # sheet's own columns and return zeros.
    SRC = "Cleaned!"

    def _rng(self, col_letter, absolute_col=False):
        c = f"${col_letter}$" if absolute_col else f"{col_letter}$"
        return f"{self.SRC}{c}{DATA_START}:{c}{self.data_end}"

    @property
    def _qrng(self):
        return (f"{self.SRC}${QUINTILE_COL}${DATA_START}"
                f":${QUINTILE_COL}${self.data_end}")

    @property
    def _rrng(self):
        return (f"{self.SRC}${REGION_COL}${DATA_START}"
                f":${REGION_COL}${self.data_end}")

    # ---------------- analysis ----------------
    def write_analysis(self):
        ws = self.wb.create_sheet("Analysis")
        ws.sheet_view.showGridLines = False
        row = 1
        ws.cell(row, 1).value = self.title
        ws.cell(row, 1).font = Font(name=FONT, bold=True, size=14)
        row += 1
        ws.cell(row, 1).value = (f"{self.n} respondents. All ranges cover rows "
                                 f"{DATA_START}-{self.data_end} of the Cleaned sheet "
                                 f"({GROWTH_PAD} blank rows reserved for growth).")
        ws.cell(row, 1).font = Font(name=FONT, italic=True, size=9)
        row += 2

        for ci, col in enumerate(self.res.columns):
            if col.kind == "free_text":
                continue
            letter = self.col_of[ci]
            if col.kind in ("coded", "multi_select_option"):
                row = self._crosstab(ws, row, col, letter)
            elif col.kind == "numeric":
                row = self._numeric(ws, row, col, letter)
            row += 2

        ws.column_dimensions["A"].width = 34
        for c in range(2, 40):
            ws.column_dimensions[get_column_letter(c)].width = 12
        return ws

    def _block_title(self, ws, row, col, letter):
        c = ws.cell(row, 1)
        c.value = col.header
        c.font = Font(name=FONT, bold=True, size=11, color="1F3864")
        ws.cell(row, 2).value = f"(Cleaned!{letter})"
        ws.cell(row, 2).font = Font(name=FONT, italic=True, size=8, color="808080")
        return row + 1

    def _crosstab(self, ws, row, col, letter):
        """Counts and row percentages by quintile and by region."""
        row = self._block_title(ws, row, col, letter)
        opts = sorted(col.codes.items(), key=lambda kv: kv[1])
        vrng = self._rng(letter)

        # option header spans two columns each (# and %), then N and Total
        hdr = row
        _label(ws.cell(hdr, 1), "", fill=SUB_FILL)
        for i, (label, code) in enumerate(opts):
            c1 = 2 + i * 2
            _label(ws.cell(hdr, c1), label, bold=True, fill=SUB_FILL)
            _label(ws.cell(hdr, c1 + 1), "", fill=SUB_FILL)
            ws.cell(hdr + 1, c1).value = "#"
            ws.cell(hdr + 1, c1 + 1).value = "%"
            for cc in (c1, c1 + 1):
                ws.cell(hdr + 1, cc).font = Font(name=FONT, bold=True, size=9)
                ws.cell(hdr + 1, cc).border = BOX
        tot_c = 2 + len(opts) * 2
        _label(ws.cell(hdr, tot_c), "Responses", bold=True, fill=SUB_FILL)
        row = hdr + 2

        for group_name, grng, members in (
            ("Quintile", self._qrng, QUINTILE_LABELS),
            ("Region", self._rrng, REGION_NAMES),
        ):
            _label(ws.cell(row, 1), group_name, bold=True, fill=TOT_FILL)
            row += 1
            first = row
            for code, name in sorted(members.items()):
                _label(ws.cell(row, 1), name)
                for i, (label, ocode) in enumerate(opts):
                    c1 = 2 + i * 2
                    cnt = ws.cell(row, c1)
                    cnt.value = f"=COUNTIFS({grng},{code},{vrng},{ocode})"
                    cnt.font = Font(name=FONT, size=10)
                    cnt.border = BOX
                    pct = ws.cell(row, c1 + 1)
                    tl = get_column_letter(tot_c)
                    pct.value = (f"=IFERROR({get_column_letter(c1)}{row}/"
                                 f"${tl}{row},\"\")")
                    pct.number_format = "0.0%"
                    pct.font = Font(name=FONT, size=10)
                    pct.border = BOX
                cells = [f"{get_column_letter(2 + i * 2)}{row}" for i in range(len(opts))]
                t = ws.cell(row, tot_c)
                t.value = "=" + "+".join(cells)
                t.font = Font(name=FONT, size=10)
                t.border = BOX
                row += 1
            last = row - 1
            _label(ws.cell(row, 1), f"TOTAL ({group_name})", bold=True, fill=TOT_FILL)
            for i in range(len(opts)):
                c1 = 2 + i * 2
                cl = get_column_letter(c1)
                s = ws.cell(row, c1)
                s.value = f"=SUM({cl}{first}:{cl}{last})"
                s.font = Font(name=FONT, bold=True, size=10)
                s.fill = TOT_FILL
                s.border = BOX
                tl = get_column_letter(tot_c)
                p = ws.cell(row, c1 + 1)
                p.value = f"=IFERROR({cl}{row}/${tl}{row},\"\")"
                p.number_format = "0.0%"
                p.font = Font(name=FONT, bold=True, size=10)
                p.fill = TOT_FILL
                p.border = BOX
            tl = get_column_letter(tot_c)
            tt = ws.cell(row, tot_c)
            tt.value = f"=SUM({tl}{first}:{tl}{last})"
            tt.font = Font(name=FONT, bold=True, size=10)
            tt.fill = TOT_FILL
            tt.border = BOX
            row += 1
        return row

    def _numeric(self, ws, row, col, letter):
        """Overall statistics plus group means by quintile and region."""
        row = self._block_title(ws, row, col, letter)
        vrng = self._rng(letter)
        vabs = self._rng(letter, absolute_col=True)

        stats = [("Responses", f"=COUNT({vabs})"), ("Mean", f"=IFERROR(AVERAGE({vabs}),\"\")"),
                 ("Median", f"=IFERROR(MEDIAN({vabs}),\"\")"),
                 ("Minimum", f"=IFERROR(MIN({vabs}),\"\")"),
                 ("Maximum", f"=IFERROR(MAX({vabs}),\"\")")]
        c = 2
        for name, f in stats:
            _label(ws.cell(row, c), name, bold=True, fill=SUB_FILL)
            cell = ws.cell(row + 1, c)
            cell.value = f
            cell.font = Font(name=FONT, size=10)
            cell.border = BOX
            c += 1
        _label(ws.cell(row, 1), "Overall", bold=True, fill=TOT_FILL)
        row += 2

        for group_name, grng, members in (
            ("Quintile", self._qrng, QUINTILE_LABELS),
            ("Region", self._rrng, REGION_NAMES),
        ):
            _label(ws.cell(row, 1), group_name, bold=True, fill=TOT_FILL)
            _label(ws.cell(row, 2), "n", bold=True, fill=SUB_FILL)
            _label(ws.cell(row, 3), "Mean", bold=True, fill=SUB_FILL)
            row += 1
            for code, name in sorted(members.items()):
                _label(ws.cell(row, 1), name)
                n = ws.cell(row, 2)
                n.value = f"=COUNTIFS({grng},{code},{vrng},\"<>\")"
                n.font = Font(name=FONT, size=10)
                n.border = BOX
                m = ws.cell(row, 3)
                m.value = f"=IF(B{row}=0,\"\",AVERAGEIF({grng},{code},{vrng}))"
                m.font = Font(name=FONT, size=10)
                m.border = BOX
                m.number_format = "#,##0.00"
                row += 1
        return row
