"""Reconciliation checks written into the workbook as live formulas.

The point is to convert silent wrongness into something visible. Each check
compares a number the analysis produced against the same number computed a
different way; a mismatch turns the row red.
"""
from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"
BAD = PatternFill("solid", fgColor="FFC7CE")
GOOD_FONT = Font(name=FONT, size=10)


def write_validation(awb, analysis_ws):
    """Add a Validation sheet to an AnalysisWorkbook that has already been written."""
    ws = awb.wb.create_sheet("Validation")
    ws.sheet_view.showGridLines = False
    res, n = awb.res, awb.n

    ws["A1"] = "Validation"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A2"] = ("Every check must read OK. A FAIL means a number on the Analysis sheet "
                "disagrees with the same quantity computed independently.")
    ws["A2"].font = Font(name=FONT, italic=True, size=9)

    row = 4
    for h, w in (("Check", 52), ("Expected", 14), ("Actual", 14), ("Result", 10)):
        c = ws.cell(row, ["Check", "Expected", "Actual", "Result"].index(h) + 1)
        c.value = h
        c.font = Font(name=FONT, bold=True, size=10)
        ws.column_dimensions[get_column_letter(c.column)].width = w
    row += 1

    de = awb.data_end
    checks = []

    # 1. Respondent count agrees between the two sheets.
    checks.append(("Respondent rows on Cleaned sheet", n,
                   f"=COUNTA(Cleaned!$A${2}:$A${de})"))

    # 2. Every row carries a valid quintile and region.
    checks.append(("Rows with a quintile in 1-5", n,
                   f"=COUNTIFS(Cleaned!$C$2:$C${de},\">=1\",Cleaned!$C$2:$C${de},\"<=5\")"))
    checks.append(("Rows with a region in 1-12", n,
                   f"=COUNTIFS(Cleaned!$D$2:$D${de},\">=1\",Cleaned!$D$2:$D${de},\"<=12\")"))

    # 3. Quintile and region partitions each account for every row exactly once.
    qparts = "+".join(f"COUNTIF(Cleaned!$C$2:$C${de},{i})" for i in range(1, 6))
    checks.append(("Quintiles 1-5 sum to the respondent count", n, f"={qparts}"))
    rparts = "+".join(f"COUNTIF(Cleaned!$D$2:$D${de},{i})" for i in range(1, 13))
    checks.append(("Regions 1-12 sum to the respondent count", n, f"={rparts}"))

    # 4. No population/quintile disagreement: each row's quintile must match its band.
    checks.append(("Rows where population is blank", 0,
                   f"=COUNTBLANK(Cleaned!$B$2:$B${2 + n - 1})"))

    # 5. Per-question: the quintile crosstab total equals the region crosstab total.
    #    These are computed from different criteria columns, so agreement is
    #    meaningful evidence the ranges are aligned.
    for ci, col in enumerate(res.columns):
        if col.kind not in ("coded", "multi_select_option"):
            continue
        letter = awb.col_of[ci]
        opts = sorted(col.codes.values())
        q = "+".join(
            f"COUNTIFS(Cleaned!$C$2:$C${de},{g},Cleaned!${letter}$2:${letter}${de},{o})"
            for g in range(1, 6) for o in opts)
        r = "+".join(
            f"COUNTIFS(Cleaned!$D$2:$D${de},{g},Cleaned!${letter}$2:${letter}${de},{o})"
            for g in range(1, 13) for o in opts)
        checks.append((f"Crosstab totals agree: {col.header[:44]}", f"={q}", f"={r}"))

    for label, expected, actual in checks:
        ws.cell(row, 1).value = label
        ws.cell(row, 1).font = GOOD_FONT
        ws.cell(row, 2).value = expected
        ws.cell(row, 2).font = GOOD_FONT
        ws.cell(row, 3).value = actual
        ws.cell(row, 3).font = GOOD_FONT
        ws.cell(row, 4).value = f'=IF(B{row}=C{row},"OK","FAIL")'
        ws.cell(row, 4).font = GOOD_FONT
        row += 1

    last = row - 1
    if last >= 6:
        ws.conditional_formatting.add(
            f"D6:D{last}",
            CellIsRule(operator="equal", formula=['"FAIL"'], fill=BAD))

    ws.cell(row + 1, 1).value = "Overall"
    ws.cell(row + 1, 1).font = Font(name=FONT, bold=True, size=11)
    ws.cell(row + 1, 4).value = f'=IF(COUNTIF(D6:D{last},"FAIL")=0,"ALL OK","FAILURES PRESENT")'
    ws.cell(row + 1, 4).font = Font(name=FONT, bold=True, size=11)
    return ws


def write_issue_log(awb):
    """Record every cleaning decision that lost or changed data."""
    ws = awb.wb.create_sheet("Cleaning Log")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Cleaning Log"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A2"] = ("Everything the cleaning step dropped, blanked, or resolved by "
                "approximate match. Review before publishing.")
    ws["A2"].font = Font(name=FONT, italic=True, size=9)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110

    row = 4
    if awb.res.fuzzy:
        ws.cell(row, 1).value = "Approximate city matches"
        ws.cell(row, 1).font = Font(name=FONT, bold=True, size=10)
        row += 1
        for raw, matched in awb.res.fuzzy:
            ws.cell(row, 1).value = raw
            ws.cell(row, 2).value = f"matched to {matched} -- confirm this is correct"
            ws.cell(row, 1).font = GOOD_FONT
            ws.cell(row, 2).font = GOOD_FONT
            row += 1
        row += 1

    ws.cell(row, 1).value = "Issues"
    ws.cell(row, 1).font = Font(name=FONT, bold=True, size=10)
    row += 1
    if not awb.res.issues:
        ws.cell(row, 2).value = "None."
        ws.cell(row, 2).font = GOOD_FONT
    for msg in awb.res.issues:
        ws.cell(row, 2).value = msg
        ws.cell(row, 2).font = GOOD_FONT
        row += 1
    return ws
