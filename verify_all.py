import sys, os, re, glob, json, subprocess, openpyxl
sys.path.insert(0,'.')
from openpyxl.utils import column_index_from_string as cidx

CT = re.compile(r"=COUNTIFS\(Cleaned!\$([CD])\$2:\$[CD]\$\d+,(\d+),Cleaned!([A-Z]{1,3})\$2:[A-Z]{1,3}\$\d+,(\d+)\)")
AV = re.compile(r'=IF\(B\d+=0,"",AVERAGEIF\(Cleaned!\$([CD])\$2:\$[CD]\$\d+,(\d+),Cleaned!([A-Z]{1,3})\$2:[A-Z]{1,3}\$\d+\)\)')

def verify(path):
    wv=openpyxl.load_workbook(path, data_only=True)
    wf=openpyxl.load_workbook(path)
    cl,anv,va = wv['Cleaned'], wv['Analysis'], wv['Validation']
    anf = wf['Analysis']
    n=0
    for r in range(2, cl.max_row+1):
        if cl.cell(r,1).value not in (None,''): n=r
    q={r:cl.cell(r,3).value for r in range(2,n+1)}
    g={r:cl.cell(r,4).value for r in range(2,n+1)}
    verdict=None; fails=0
    for r in range(6, va.max_row+1):
        if va.cell(r,4).value=='FAIL': fails+=1
        if va.cell(r,1).value=='Overall': verdict=va.cell(r,4).value
    nct=bad_ct=nav=0; worst=0.0
    for row in anf.iter_rows():
        for c in row:
            v=c.value
            if not isinstance(v,str): continue
            m=CT.fullmatch(v)
            if m:
                grp = q if m.group(1)=='C' else g
                vi=cidx(m.group(3)); gv=int(m.group(2)); ov=int(m.group(4))
                exp=sum(1 for r in range(2,n+1) if grp.get(r)==gv and cl.cell(r,vi).value==ov)
                nct+=1
                if anv[c.coordinate].value!=exp: bad_ct+=1
                continue
            m=AV.fullmatch(v)
            if m:
                grp = q if m.group(1)=='C' else g
                vi=cidx(m.group(3)); gv=int(m.group(2))
                vals=[cl.cell(r,vi).value for r in range(2,n+1)
                      if grp.get(r)==gv and isinstance(cl.cell(r,vi).value,(int,float))]
                got=anv[c.coordinate].value; nav+=1
                if vals and got is not None:
                    exp=sum(vals)/len(vals)
                    rel=abs(got-exp)/max(abs(exp),1e-12)
                    worst=max(worst,rel)
    return dict(rows=n-1, verdict=verdict, val_fails=fails, countifs=nct,
                countifs_bad=bad_ct, averageif=nav, max_rel_diff=worst)

results={}
for p in sorted(glob.glob('output/*.analysis.xlsx')):
    name=os.path.basename(p).replace('.analysis.xlsx','')[:42]
    r=subprocess.run([sys.executable,'/mnt/skills/public/xlsx/scripts/recalc.py',p,'240'],
                     capture_output=True,text=True)
    try: rep=json.loads(r.stdout)
    except Exception: rep={'status':'?','total_formulas':0,'total_errors':-1}
    v=verify(p)
    results[name]=dict(**v, formulas=rep.get('total_formulas'), errors=rep.get('total_errors'))
    print(f"{name:44s} formulas={rep.get('total_formulas'):>6,} errors={rep.get('total_errors'):>2} "
          f"| validation={v['verdict']:>6} fails={v['val_fails']:>2} "
          f"| COUNTIFS {v['countifs']:>5,} bad={v['countifs_bad']} "
          f"| AVERAGEIF {v['averageif']:>4,} maxdiff={v['max_rel_diff']:.1e}")
json.dump(results, open('output/_verification.json','w'), indent=2)
